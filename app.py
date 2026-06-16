"""
시간표 자동 배정 웹앱 (FastAPI)
실행(로컬): python app.py  ->  http://localhost:8000

데이터 저장:
  - 기본: SQLite (앱 폴더의 schedule.db)
  - 환경변수 DATABASE_URL 이 있으면 그 DB(예: Render Postgres)에 영구 저장
"""
import os, io, json, secrets, datetime, urllib.parse
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List
from sqlalchemy import (create_engine, MetaData, Table, Column, String, Integer,
                        Text, select, insert, update, delete)

import solver as solver_mod

BASE = os.path.dirname(os.path.abspath(__file__))
# HTML 파일은 app.py 와 같은 폴더에 둔다 (static 하위폴더 없이도 동작)
STATIC = os.path.join(BASE, "static") if os.path.isdir(os.path.join(BASE, "static")) else BASE

# ---------- DB (SQLite 로컬 / Postgres 클라우드 공용) ----------
DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL:
    # Render 등은 postgres:// 로 주는데 SQLAlchemy는 postgresql+psycopg:// 필요
    url = DATABASE_URL
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql+psycopg://", 1)
    elif url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)
    engine = create_engine(url, pool_pre_ping=True)
else:
    DB = os.environ.get("SCHEDULE_DB", os.path.join(BASE, "schedule.db"))
    engine = create_engine(f"sqlite:///{DB}")

meta = MetaData()
events = Table("events", meta,
               Column("id", String(64), primary_key=True),
               Column("admin_token", String(64)),
               Column("title", String(255)),
               Column("config", Text),
               Column("result", Text),
               Column("created_at", String(64)))
participants = Table("participants", meta,
                     Column("id", Integer, primary_key=True, autoincrement=True),
                     Column("event_id", String(64)),
                     Column("name", String(255)),
                     Column("slots", Text),
                     Column("created_at", String(64)))
meta.create_all(engine)

app = FastAPI(title="시간표 자동 배정")


def now():
    return datetime.datetime.now().isoformat()


# ---------- helpers ----------
def hours_to_slots(hours, slot_minutes):
    return max(1, round(hours * 60 / slot_minutes))


def build_times(day_start, day_end, slot_minutes):
    sh, sm = map(int, day_start.split(":"))
    eh, em = map(int, day_end.split(":"))
    start, end = sh * 60 + sm, eh * 60 + em
    times = []
    t = start
    while t + slot_minutes <= end:
        a, b = t, t + slot_minutes
        times.append({"label": f"{a//60:02d}:{a%60:02d}~{b//60:02d}:{b%60:02d}", "start": a, "end": b})
        t += slot_minutes
    return times


def place_rows(place, times):
    def mm(x):
        h, m = map(int, x.split(":"))
        return h * 60 + m
    a, b = mm(place["op_start"]), mm(place["op_end"])
    return [i for i, t in enumerate(times) if t["start"] >= a and t["end"] <= b]


# ---------- models ----------
class PlaceIn(BaseModel):
    name: str
    cap: int = 1
    priority: int = 1
    op_start: str = "09:00"
    op_end: str = "21:00"
    required: bool = False


class EventIn(BaseModel):
    title: str = "근무 시간표"
    days: List[str] = ["월", "화", "수", "목", "금"]
    day_start: str = "09:00"
    day_end: str = "21:00"
    slot_minutes: int = 60
    min_run_hours: float = 3
    max_day_hours: float = 8
    places: List[PlaceIn]


class ParticipantIn(BaseModel):
    name: str
    slots: List[int]


# ---------- API ----------
@app.post("/api/events")
def create_event(ev: EventIn):
    times = build_times(ev.day_start, ev.day_end, ev.slot_minutes)
    if not times:
        raise HTTPException(400, "시간 범위가 잘못됐어요 (종료가 시작보다 빨라요).")
    places = []
    for pl in ev.places:
        d = pl.dict()
        d["rows"] = place_rows(d, times)
        if not d["rows"]:
            raise HTTPException(400, f"'{pl.name}' 운영시간이 전체 시간과 안 맞아요.")
        places.append(d)
    config = {
        "title": ev.title, "days": ev.days, "day_start": ev.day_start, "day_end": ev.day_end,
        "slot_minutes": ev.slot_minutes, "times": times, "places": places,
        "min_run_slots": hours_to_slots(ev.min_run_hours, ev.slot_minutes),
        "max_day_slots": hours_to_slots(ev.max_day_hours, ev.slot_minutes),
        "min_run_hours": ev.min_run_hours, "max_day_hours": ev.max_day_hours,
    }
    eid = secrets.token_urlsafe(6)
    token = secrets.token_urlsafe(8)
    with engine.begin() as conn:
        conn.execute(insert(events).values(
            id=eid, admin_token=token, title=ev.title,
            config=json.dumps(config, ensure_ascii=False), result=None, created_at=now()))
    return {"event_id": eid, "admin_token": token}


@app.get("/api/events/{eid}")
def get_event(eid: str):
    with engine.connect() as conn:
        row = conn.execute(select(events).where(events.c.id == eid)).mappings().first()
    if not row:
        raise HTTPException(404, "이벤트를 찾을 수 없어요.")
    return {"event_id": eid, "title": row["title"], "config": json.loads(row["config"])}


@app.post("/api/events/{eid}/participants")
def add_participant(eid: str, p: ParticipantIn):
    with engine.begin() as conn:
        ev = conn.execute(select(events.c.id).where(events.c.id == eid)).first()
        if not ev:
            raise HTTPException(404, "이벤트를 찾을 수 없어요.")
        name = p.name.strip()
        if not name:
            raise HTTPException(400, "이름을 입력해주세요.")
        conn.execute(delete(participants).where(
            (participants.c.event_id == eid) & (participants.c.name == name)))
        conn.execute(insert(participants).values(
            event_id=eid, name=name, slots=json.dumps(p.slots), created_at=now()))
    return {"ok": True}


@app.get("/api/events/{eid}/participants")
def list_participants(eid: str):
    with engine.connect() as conn:
        rows = conn.execute(select(participants).where(
            participants.c.event_id == eid).order_by(participants.c.id)).mappings().all()
    return [{"name": r["name"], "slots": json.loads(r["slots"]), "created_at": r["created_at"]} for r in rows]


@app.post("/api/events/{eid}/solve")
def run_solve(eid: str, token: str = ""):
    with engine.connect() as conn:
        row = conn.execute(select(events).where(events.c.id == eid)).mappings().first()
        if not row:
            raise HTTPException(404, "이벤트를 찾을 수 없어요.")
        if token != row["admin_token"]:
            raise HTTPException(403, "관리자만 배정을 실행할 수 있어요.")
        config = json.loads(row["config"])
        prows = conn.execute(select(participants).where(
            participants.c.event_id == eid)).mappings().all()
    parts = [{"name": r["name"], "slots": json.loads(r["slots"])} for r in prows]
    if not parts:
        raise HTTPException(400, "제출한 참가자가 없어요.")
    result = solver_mod.solve(config, parts, time_limit=int(os.environ.get("SOLVE_TIME", "30")))
    with engine.begin() as conn:
        conn.execute(update(events).where(events.c.id == eid).values(
            result=json.dumps(result, ensure_ascii=False)))
    return result


@app.get("/api/events/{eid}/result")
def get_result(eid: str):
    with engine.connect() as conn:
        row = conn.execute(select(events.c.result).where(events.c.id == eid)).first()
    if not row:
        raise HTTPException(404, "이벤트를 찾을 수 없어요.")
    if not row[0]:
        return {"solved": False}
    return {"solved": True, "result": json.loads(row[0])}


def build_xlsx(config, result):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    days, times, places = config["days"], config["times"], config["places"]
    nT = len(times)
    slotH = config["slot_minutes"] / 60
    assign, loads = result["assign"], result["loads"]
    blue = PatternFill("solid", fgColor="9DC3E6"); gray = PatternFill("solid", fgColor="D9D9D9")
    green = PatternFill("solid", fgColor="C6E0B4"); red = PatternFill("solid", fgColor="F4B7B7")
    dark = PatternFill("solid", fgColor="808080")
    bold = Font(bold=True); ctr = Alignment(horizontal="center", vertical="center")
    thin = Border(*[Side(style="thin", color="BBBBBB")] * 4)
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "배정결과"
    ws["A1"] = config.get("title", "배정 결과"); ws["A1"].font = Font(bold=True, size=13)
    r = 3
    for di, dn in enumerate(days):
        ws.cell(r, 1, dn).font = Font(bold=True, size=12); ws.cell(r, 1).fill = blue
        for pi, pl in enumerate(places):
            c = ws.cell(r, 2 + pi, f'{pl["name"]}({pl["cap"]})'); c.font = bold; c.fill = blue; c.alignment = ctr; c.border = thin
        r += 1
        for ti, t in enumerate(times):
            s = di * nT + ti
            ws.cell(r, 1, t["label"]).fill = gray; ws.cell(r, 1).border = thin
            for pi, pl in enumerate(places):
                cell = ws.cell(r, 2 + pi); cell.alignment = ctr; cell.border = thin
                if ti not in pl["rows"]:
                    cell.value = "운영 안함"; cell.fill = dark; cell.font = Font(color="FFFFFF")
                else:
                    who = assign.get(f"{s}_{pi}", [])
                    cell.value = ", ".join(who) if who else "·"
                    cell.fill = green if who else red
            r += 1
        r += 1
    ws.cell(r, 1, "이름").font = bold
    for di, dn in enumerate(days):
        c = ws.cell(r, 2 + di, dn); c.font = bold; c.fill = blue; c.alignment = ctr
    ws.cell(r, 2 + len(days), "총(시간)").font = bold; ws.cell(r, 2 + len(days)).fill = blue
    r += 1
    for name in loads:
        ws.cell(r, 1, name).font = bold
        for di in range(len(days)):
            cnt = sum(1 for ti in range(nT) for pi in range(len(places))
                      if name in assign.get(f"{di*nT+ti}_{pi}", []))
            ws.cell(r, 2 + di, round(cnt * slotH, 2) or "").alignment = ctr
        ws.cell(r, 2 + len(days), round(loads[name] * slotH, 2)).font = bold
        r += 1
    ws.column_dimensions["A"].width = 13
    for col in "BCDEFG": ws.column_dimensions[col].width = 18
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


@app.get("/api/events/{eid}/export.xlsx")
def export_xlsx(eid: str):
    with engine.connect() as conn:
        row = conn.execute(select(events.c.title, events.c.config, events.c.result)
                           .where(events.c.id == eid)).mappings().first()
    if not row:
        raise HTTPException(404, "이벤트를 찾을 수 없어요.")
    if not row["result"]:
        raise HTTPException(400, "아직 배정을 실행하지 않았어요.")
    data = build_xlsx(json.loads(row["config"]), json.loads(row["result"]))
    quoted = urllib.parse.quote(f"{row['title']}_배정결과.xlsx")
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted}"})


# ---------- 페이지 ----------
@app.get("/")
def page_home():
    return FileResponse(os.path.join(STATIC, "index.html"))


@app.get("/e/{eid}")
def page_participant(eid: str):
    return FileResponse(os.path.join(STATIC, "participant.html"))


@app.get("/admin/{eid}")
def page_admin(eid: str):
    return FileResponse(os.path.join(STATIC, "admin.html"))


app.mount("/static", StaticFiles(directory=STATIC), name="static")


@app.get("/health")
def health():
    return {"ok": True}


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    print(f"\n  시간표 배정 서버 시작!  ->  http://localhost:{port}\n")
    uvicorn.run(app, host="0.0.0.0", port=port)
